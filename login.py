import random
from tkinter import *
#import captchaLib
import random
import string
from captcha.image import ImageCaptcha
import openpyxl
from openpyxl import load_workbook


root = Tk()  # main 설정
root.title("log in to team10")  # 위에 뜨는 거
root.geometry("400x400+100+100")  # 넓이 설정
root.resizable(False, False)  # 크기 조정 가능한지 x축, y축

MAX_LENGTH = 6

title = Label(root, text="Log In", font='Helvetica 15 bold')
# root는 위치 할 곳, text는 표시할 문자,font는 설정
title.pack(pady='10')
# pack은 가운데에 위치 시키는 거고, pady는 padding y축 (y축으로 위아래 10씩!)

mainFrame = Frame(root)
# Frame은 contents가 들어갈 곳! root에 위치시켜도 되는데
mainFrame.pack()

# email을 저장할 frame(emailFrame)을 main frame에 위치 시킴
emailFrame = Frame(mainFrame)
emailFrame.pack()
lblEmail = Label(emailFrame, text="email", width='8', anchor="w")  # 넓이 8, anchor='w'해서 왼쪽 정렬
lblEmail.grid(row=1, column=0)  # grid는 격자로 해서 위치 지정!
email = Entry(emailFrame)
email.grid(row=1, column=1)

# password을 저장할 frame(emailFrame)을 main frame에 위치 시킴
passwordFrame = Frame(mainFrame)
passwordFrame.pack()
lblPassword = Label(passwordFrame, text="password", width='8', anchor="w")
lblPassword.grid(row=1, column=0)
password = Entry(passwordFrame, show = '*')
password.grid(row=1, column=1)

def print_random():
    _LENGTH_=6
    string_pool = string.ascii_uppercase + string.digits

    result = ""
    for i in range(_LENGTH_):
        result += random.choice(string_pool)
    global tmp_str
    tmp_str = result
    print(tmp_str)
    return result


#보안문자
randomframe = Frame(mainFrame)

#새로고침
def refresh_btncmd():
    global randomframe
    randomframe.pack()

    image = ImageCaptcha(width=160, height=90)
    txt_captcha = print_random()
    image.generate(txt_captcha)
    image.write(txt_captcha, 'captcha_result.png')
    photo = PhotoImage(file='captcha_result.png')

    labelrandom = Label(randomframe, image=photo)
    labelrandom.grid(row = 1, column=0)
    global rand
    rand = Entry(randomframe, width = 15, font = 8)
    rand.place(x = 40, y = 20)
    rand.grid(row = 2)
refresh_btncmd()

ref_btn = Button(randomframe, text = "새로고침", anchor = "e", command = refresh_btncmd)
ref_btn.grid(row = 1, column = 1)

count = 0


def login():
    global count

    fpath = r'C:\pystudy\test_data.xlsx' #db파일이 저장될 경로는 각자 수정하기
    wb= openpyxl.load_workbook(fpath)
    ws = wb.active
    read_xlsx = wb
    
    #db에 있는 이메일 목록 추출
    read_sheet = read_xlsx.active
    name_col = read_sheet['C']
    dbemail = []
    for cell in name_col:
        dbemail.append(cell.value)

    #입력한 이메일과 동일한 이메일이 db에 있는지 확인
    if email.get() in dbemail:
        k = dbemail.index(email.get())
        dbpassword = ws[f'D{k+1}'].value
        


    if (email.get() == dbemail[k]) and (password.get() == dbpassword) and (rand.get() == tmp_str):
        new = Tk()
        new.title("successful login")
        new.geometry("400x400+100+100")
        new.resizable(False, False)
        welcome = Label(new, text="Welcome!!", width=20, height=50)
        welcome.pack()
        new.mainloop()
    else:
        count += 1
        notice2.config(text=str(count) + "회 실패하셨습니다")


notice2 = Label(root)
notice2.pack()

loginbuttonFrame = Frame(mainFrame)
loginbuttonFrame.pack(pady=(10, 20))
loginbutton = Button(loginbuttonFrame, overrelief="solid", width=20, command=login, repeatdelay=1000,
                     repeatinterval=100, text="로그인")
loginbutton.grid(row=1, column=1)


#회원가입창으로 이동
def new_window():
    root.destroy()
    import signin #회원가입 파일명

#회원가입 버튼
signinbuttonFrame = Frame(mainFrame)
signinbuttonFrame.pack(pady=(10, 20))
signinbutton = Button(signinbuttonFrame, overrelief="solid", width=20, command=new_window, repeatdelay=1000,
                     repeatinterval=100, text="회원가입")
signinbutton.grid(row=1, column=1)

root.mainloop()
