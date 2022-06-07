from glob import glob
from tkinter import *
import tkinter.messagebox as msgbox
import random
import string
from captcha.image import ImageCaptcha
import sys
import openpyxl
import os
import captchaLib

fpath=r'C:\Users\82103\Desktop\pystudy\test_data.xlsx'


# 회원가입시  db에 저장되게 하는 함수
def save_info(sex, name, email, password):
    # 기존에 db 파일이 없는 경우
    if os.path.exists(fpath) == FALSE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '회원정보'
        ws['A1'] = '성별'
        ws['B1'] = '이름'
        ws['C1'] = '이메일'
        ws['D1'] = '비밀번호'
        row = 2

        if sex == 1:
            ws.cell(row, 1).value = "남"

        else:
            ws.cell(row, 1).value = "여"

        ws.cell(row, 2).value = name
        ws.cell(row, 3).value = email
        ws.cell(row, 4).value = password

    # 기존에 db파일이 이미 있는 경우
    else:
        wb = openpyxl.load_workbook(fpath)
        ws = wb.active
        row = ws.max_row + 1

        if sex == 1:
            ws.cell(row, 1).value = "남"

        else:
            ws.cell(row, 1).value = "여"

        ws.cell(row, 2).value = name
        ws.cell(row, 3).value = email
        ws.cell(row, 4).value = password

    wb.save(fpath)


root = Tk()
root.title("Sign in to team10")

title = Label(root, text="CLOUD 회원가입", font=('Helvetica 20 bold'))
root.geometry("400x400+200+100")  # 윈도우/창이 나타나는 위치를 지정
title.pack()

mainFrame = Frame(root)
mainFrame.pack()

sexframe = Frame(mainFrame)
sexframe.pack()
labelsex = Label(sexframe, text="성별", width=16, font=("맑은 고딕", '9', 'bold'), anchor="w")
labelsex.grid(row=1, column=0)

sex = IntVar()
male = Radiobutton(sexframe, text="남성", anchor="e", variable=sex, value=1)
female = Radiobutton(sexframe, text="여성", anchor="e", variable=sex, value=2)
male.grid(row=1, column=1)
female.grid(row=1, column=2)

nameframe = Frame(mainFrame)
nameframe.pack()
labelname = Label(nameframe, text="이름", width=10, font=("맑은 고딕", '9', 'bold'), anchor="w")
labelname.grid(row=1, column=0)
name = Entry(nameframe)
name.grid(row=1, column=1)

emailframe = Frame(mainFrame)
emailframe.pack()
labelemail = Label(emailframe, text="e-mail", width=10, font=("맑은 고딕", '9', 'bold'), anchor="w")
labelemail.grid(row=1, column=0)
email = Entry(emailframe)
email.grid(row=1, column=1)
email_str = email.get()


# 이메일 형식 판독 함수
def is_email_valid(email_str):
    # 이메일 형식이 알맞지 않은 경우
    if email_str.rfind("@") == -1 or email_str.rfind(".") == -1:
        return FALSE
    # 이메일 형식이 알맞은 경우
    else:
        return TRUE


pwframe = Frame(mainFrame)
pwframe.pack()
labelpw = Label(pwframe, text="password", width=10, font=("맑은 고딕", '9', 'bold'), anchor="w")
labelpw.grid(row=1, column=0)
pw = Entry(pwframe, show='*')
pw.grid(row=1, column=1)

# 보안문자
randomframe = Frame(mainFrame)


# 새로고침
def refresh_btncmd():
    global tmp_str, randomframe
    randomframe.pack()
    tmp_str, photo = captchaLib.refresh()
    labelrandom = Label(randomframe, image=photo)
    labelrandom.grid(row=1, column=0)
    global rand
    rand = Entry(randomframe, width=15, font=8)
    rand.place(x=40, y=20)
    rand.grid(row=2)


refresh_btncmd()

ref_btn = Button(randomframe, text="새로고침", anchor="e", command=refresh_btncmd)
ref_btn.grid(row=1, column=1)


# 등록 버튼 기능
def signin_btncmd():
    # 입력되지 않은 정보가 있는 경우
    if name.get() == "" or email.get() == "" or pw.get() == "" or (sex.get() != 1 and sex.get() != 2):
        msgbox.showwarning("오류", "입력되지 않은 정보가 있습니다. 다시 확인하세요.")

    else:
        # 이메일 형식이 알맞지 않은 경우
        if is_email_valid(email.get()) == FALSE:
            msgbox.showwarning("오류", "이메일 형식이 알맞지 않습니다. 형식을 다시 확인하세요.")
        # 이메일 형식이 알맞은 경우
        else:
            if tmp_str == rand.get().upper():
                save_info(sex.get(), name.get(), email.get(), pw.get())
                msgbox.showinfo("완료", "회원가입이 완료되었습니다")
                root.destroy()  # 창 닫기
                import Login  # 회원가입 완료 후 로그인 창으로 이동

            # 보안문자가 일치하지 않는 경우
            else:
                msgbox.showwarning("오류", "보안문자가 일치하지 않습니다. 다시 입력하세요.")


# 등록 버튼
signin_btn = Button(root, text="등록", anchor=CENTER, command=signin_btncmd, width=10, height=2)
signin_btn.pack()

root.mainloop()
